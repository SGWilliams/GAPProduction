## This module facilitates common tasks for searching/manipulating plain text
## tables.
##
##
## The public functions are:
##
## DetermineDelimiter() - Attempts to determine the delimiter used in a table
##      file.
##
## DictionaryOfCells() - Creates a dictionary in which the keys are tuples of
##      the row header and the column header and the value is the value found at
##      the cell identified by that row/column intersection.
##
## DictionaryFromColumns() - Creates a dictionary with pairs from any two
##      selected fields in a table.
##
## DictionaryFromTable() - Creates a dictionary for which the key is the content
##      from the selected column, and the value is a tuple of the content from
##      all other columns in the table.
##
## ListHeaders() - Returns a list containing the table's field headers/column
##      names
##
## ListValues() -- Returns a list of all values that occur in the given field.
##
## ListFromTable() - Creates a list of lists from the passed table.
##
## WriteSppTable() - Creates a table (including headings) with the taxonomic and
##      crosswalking info as well as the GAP range/model status for all the
##      input species.
##
## WriteListToCsv() - Writes the passed list to a csv table
##
## WriteDictToTable() - Writes the passed dictionary to a csv table
##
## AttributeTableToCsv() - Exports the attribute table from the passed
##      shapefile, feature class, raster, or geodatabase table to a csv.
##

import os, csv
import gapdb, docs, featureclasses


###################################################
############### LISTING VALUES ####################
###################################################

def ListValues(table, field, unique=False):
    '''
    (str, str, [boolean]) -> list

    Returns a list of all values for the given field within the given table.
        Operates on csv, txt, xlsx, xls, and shp files, as well as geodatabase
        feature classes.

        Note: This function assumes that the first row contains field headings.

    Arguments:
    table -- Name/path of the file from which you wish to identify values
    field -- Name of the field/column from which you wish to identify values
    unique -- An optional boolean, indicating whether you wish to remove\
        repeated values. By default, it is set to False, indicating that
        duplicates will not be reduced.
    '''
    # Identify the file type and call the appropriate function for the type
    if table.endswith('.csv'):
        l = __ListValues_csv(table, field, unique)
    elif table.endswith('txt'):
        l = __ListValues_txt(table, field, unique)
    elif table.endswith('.xlsx') or table.endswith('.xls'):
        l = __ListValues_excel(table, field, unique)
    #elif table.endswith('.shp') or table.endswith('.dbf') or '.gdb' in os.path.abspath(table):
    else:
        try:
            l = featureclasses.ListValues(table, field, unique)
        except Exception as e:
            print e

    return l


def __ListValues_txt(table, field, unique):
    # Get a list of the rows
    t = ListFromTable(table)
    # The first row (headers)
    headers = t[0]
    # Set the index as that matching the field name
    ind = headers.index(field)
    # Instantiate an empty list
    l = list()
    # For each subsequent row in the table
    for row in t[1:]:
        # Append that row's value from the given index
        l.append(row[ind])

    if unique:
        l = list(set(l))

    return l



# Function to list the values for a given field from an Excel file
def __ListValues_excel(table, field, unique):
    # Attempt to import the necesssary library
    try:
        import openpyxl
    except:
        print 'To perform the function ListValues() on an Excel file, you must have the openpyxl library installed.'
        return False

    # If the passed field is an integer, assume it is the index
    if type(field) == int:
        ind = field
    # Otherwise, assume the user passed the field name
    else:
        # Call the function to get the index of the field name
        ind = GetColumnIndex_excel(table, field)

    # Instantiate a list to store the values
    l = list()

    # Create the workbook and worksheet objects
    wb = openpyxl.load_workbook(filename = table, use_iterators = True)
    wsName = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(name = wsName)

    # Instantiate a row count (will be used to skip the first row)
    count = 0
    # For each row in the worksheet
    for row in ws.iter_rows():
        # If the row count is less than one
        if count < 1:
            # Increment the row count and skip to the next row
            count += 1
            continue
        # Append that row's value from the given column index
        l.append(row[ind].internal_value)

    # If the user requested to return only unique values, remove duplicates
    if unique:
        l = list(set(l))

    return l


# Identify the index for the passed field/column name in an Excel file
def GetColumnIndex_excel(table, field):
    # Attempt to import the necessary library
    try:
        import openpyxl
    except:
        print 'To perform the function GetColumnIndex_excel(), you must have the openpyxl library installed.'

    # Set the index to -1
    i = -1

    headers = ListHeaders(table)
    for ind, h in enumerate(headers):
        if h == field:
            i = ind
            break

    # If the index was not updated, the field must not exist, so return False
    if i < 0:
        i = False
    return i


#def GetColumnIndex_excel(table, field):
#    # Attempt to import the necessary library
#    try:
#        import openpyxl
#    except:
#        print 'To perform the function GetColumnIndex_excel(), you must have the openpyxl library installed.'
#
#    # Open the workbook and worksheet
#    wb = openpyxl.load_workbook(filename = table, use_iterators = True)
#    wsName = wb.get_sheet_names()[0]
#    ws = wb.get_sheet_by_name(name = wsName)
#
#    # Instantiate a row counter
#    count = int()
#    # Set the index to -1
#    i = -1
#    # For each row
#    for row in ws.iter_rows():
#        # If the row counter is greater than zero, break out of the loop
#        if count > 0:
#            break
#        # For each index and object in the row
#        for j, k in enumerate(row):
#            # Get the cell's value
#            n = k.internal_value
#            print n
#            # If the cell's value is the same as the passed field name
#            if n == field:
#                print '\tSuccess!!'
#                # Set the index to that cell's column index
#                i = j
#                # End the loop
#                break
#        # Increment the counter
#        count += 1
#    # If the index was not updated, the field must not exist, so return False
#    if i < 0:
#        i = False
#    return i


# Function to list the values for a given field from a csv file
def __ListValues_csv(table, field, unique):
    import csv
    # Instatiate a list to store the records
    l = list()
    # Read the csv
    with open(table, 'rb') as csvF:
        csvReader = csv.reader(csvF)
        # If the user passed an integer for the field
        if type(field) == int:
            for row in csvReader:
                # Use the passed index to find the proper value and append it
                l.append(row[field])
        # Otherwise, assume that the user passed the field header
        else:
            # Read the first line and assign to headers
            headers = csvReader.next()
            # Get the index of the passed field name from the headers list
            i = headers.index(field)
            # For each subsequent row
            for row in csvReader:
                # Append the item from the determined index
                l.append(row[i])
    # If the user wishes to see only unique values
    if unique:
        # Remove duplicates from the list
        l = list(set(l))

    return l






###################################################
################### HEADERS #######################
###################################################

def ListHeaders(table):
    '''
    (str) -> list

    Returns a list of the table's field headings/column names.

    Argument:
    table - The file from which you wish to identify the headers.

    Example:
    >>> ListHeaders('testTable.csv')
    ['UID','SpCode','Common_Name']
    '''
    # Identify the file type and call the appropriate function for the type
    if table.endswith('.csv'):
        l = __ListHeaders_csv(table)
    elif table.endswith('txt'):
        l = __ListHeaders_txt(table)
    elif table.endswith('.xlsx') or table.endswith('.xls'):
        l = __ListHeaders_excel(table)
    elif table.endswith('.shp') or '.gdb' in os.path.abspath(table):
        l = featureclasses.ListFields(table)

    return l


def __ListHeaders_txt(table):
    delimiter = DetermineDelimiter(document)

    # If no legitimate delimiter was found, return None
    if delimiter is False:
        print "Could not determine the delimiter used for %s." % document
        return False

    with open(table, 'r') as content:
        headers = content.readline().split(delimiter)
    return headers


def __ListHeaders_csv(table):
    import csv
    with open(table, 'rb') as csvF:
        csvReader = csv.reader(csvF)
        # Read the first row
        headers = csvReader.next()
    return headers


def __ListHeaders_excel(table):
    # Attempt to import the necessary library
    try:
        import openpyxl
    except:
        print 'To perform the function ListHeaders_excel(), you must have the openpyxl library installed.'

    wb = openpyxl.load_workbook(filename = table)
    wsName = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(name = wsName)

    l = list()
    for x in range(1, ws.get_highest_column() + 1):
        col = openpyxl.cell.get_column_letter(x)
        coord = '{0}1'.format(col)
        n = ws.cell(coord).value
        l.append(n)

    return l



###################################################
################# DELIMITERS ######################
###################################################

##################################
#### Private function to check whether the test delimiter is valid
def __CheckDelimiter(document, delm):
    '''
    Private function that should only be called by other functions within this
    module.
    '''
    # Open the document
    cont = open(document, 'r')

    try:
        # Get the length of the first row split by the test delimiter
        lenRow1 = len(cont.readline().strip().split(delm))
        # Reset the cursor to the first row
        cont.seek(0)

        # Read all lines in the document
        for line in cont.readlines():
            # If the delimiter is not in a line, return False
            if delm not in line:
                return False
            # If any line does not hold the same number of records as the first
            # line, as split by the test delimiter, return False
            fields = line.strip().split(delm)
            if len(fields) != lenRow1:
                return False
        # If every line holds the same number of records, as split by the
        # test delimiter, return the delimiter
        return True

    except Exception as e:
        print e

    # Close the document on the way out of the function
    finally:
        cont.close()


##################################
#### Public function to determine the delimiter of a given table file
def DetermineDelimiter(document):
    """
    (string) -> string/boolean

    Returns the delimiter used in the document. Returns False if the delimiter
    cannot be deciphered.

    Note that it is possible for a document to be structured in such a way that
    it appears to be delimited by a character that was not intended as the
    delimiter (just open any document into Microsoft Excel, and you'll see that
    programmers far greater than me have still not resolved this issue.

    This function will only identify tabs, commas, semi-colons, colons, spaces,
    double dashes, or a single dash as the delimiters, and it will search in
    that order.

    Example:
    >>> DetermineDelimiter('test.csv')
    ,
    """
    # If the file does not exist, return False
    if os.path.exists(document) is False:
        print 'The document that you provided to tables.DetermineDelimiter(),\n%s, does not exist.' % document
        return False

    # The list of delimiters to test
    delms = ['\t', ',', ';', ':', ' ', '--', '-']

    # For each delimiter in the list
    for delm in delms:
        # Call the function to check the delimiter's validity; if the function
        # returns True, then return the delimiter
        if __CheckDelimiter(document, delm) == True:
            return delm
    # If no valid delimiter was found, return False
    return False




###################################################
################# DICTIONARIES ####################
###################################################

def DictionaryOfCells(table, rowHeaderColumn=0):
    '''
    (str) -> dictionary

    Creates a dictionary in which the keys are tuples of the row header and the
      column header and the value is the value found at the cell identified by
      that row/column intersection.

      Note that the meaningfulness of the output dictionary relies on the first
      row and the first column (or the rowHeaderColumn) in the table containing
      unique identifiers/field names.

    Arguments:
    table - The path/name of the table from which you would like to create a
      dictionary.
    rowHeaderColumn - The column header or the zero-based index of the column
      from which you wish to extract row headers. That is, if the first column
      in your table is an OID, but for easier interpretation, you'd prefer to
      use a user-assigned field, named "PointID", you could pass either "PointID"
      or the index of the column (i.e., if it's the last column, pass -1; if it's
      the third column, pass 2). Values from this column will then occur in the
      dictionary keys (along with column headers).

    Example:
    given a table (t) that looks like:
    PointID   State   SurveyCount
    293       IN      3
    495       IL      46
    >>> DictionaryOfCells(t)
    {('495', 'State'): 'IL', ('293', 'SurveyCount'): '3', ('293', 'State'): 'IN', ('495', 'SurveyCount'): '46'}
    >>> gap.tables.DictionaryOfCells(t, 'State')
    {('IN', 'SurveyCount'): '3', ('IL', 'PointID'): '495', ('IN', 'PointID'): '293', ('IL', 'SurveyCount'): '46'}
    '''
    rows = ListFromTable(table)
    headers = rows[0]

    d = dict()

    if not type(rowHeaderColumn) is int:
        rowHeaderColumn = headers.index(rowHeaderColumn)

    indices = [i for i in range(len(headers)) if i != rowHeaderColumn]
    indices.sort()

    for row in rows[1:]:
        for x in indices:
            d[(row[rowHeaderColumn], headers[x])] = row[x]

    return d



# Public function to create a dictionary from the passed table and the passed
# key and value fields
def DictionaryFromColumns(table, keyColumn, valueColumn, includeHeaders=True):
    '''
    (string, string/integer, string/integer, [string], [string]) -> dictionary

    Returns a dictionary of the values from the selected fields in a table.

    Note: This function is to return a dictionary in which the value represents
    a single column from the table. If you want the dictionary's values to hold
    all columns from the table (i.e., the value is a tuple), use the
    DictionaryFromTable() function (though that function currently only handles
    txt files).

    Also, a float cannot be used as a dictionary key; therefore, if the user
    passes a field containing floats as the key field, the floats will be
    converted to strings in the dictionary.

    Further, if the passed keyColumn does not contain unique values in every
    row, not all relevant data from the table will be contained in the
    output dictionary.

    Works on csv, txt, Excel, or ArcGIS files.

    Arguments:
    document -- The absolute path to the table file.
    keyColumn -- The header text or the 0-based index of the column that will be
        assigned as the dictionary key.
    valueColumn -- The header text or the 0-based index of the column that will
        be assigned as the dictionary value.

    Examples:
    >>> DictionaryFromColumns('test.csv', 'eng', 3, 'o')
    {'four': 'quatro', 'two': 'dos', 'one': 'uno'}
    >>> DictionaryFromColumns('test.txt', 1, 'span')
    {'num': 'span', '1': 'uno', '3': 'tres', '2': 'dos', '5': 'cinco', '4': 'quatro'}
    '''
    # Identify the file type and call the appropriate function for the type
    if table.endswith('.csv'):
        d = __DictionaryFromColumns_csv(table, keyColumn, valueColumn, includeHeaders)
    elif table.endswith('txt'):
        d = __DictionaryFromColumns_txt(table, keyColumn, valueColumn, includeHeaders)
    elif table.endswith('.xlsx') or table.endswith('.xls'):
        d = __DictionaryFromColumns_excel(table, keyColumn, valueColumn)
    #elif table.lower().endswith('.shp') or '.gdb' in os.path.abspath(table) or table.lower().endswith('.dbf'):
    else:
        d = featureclasses.FieldsDictionary(table, keyColumn, valueColumn)

    return d


# Private function to create a dictionary from the passed table and the passed
# key and value fields for a csv file
def __DictionaryFromColumns_csv(table, keyColumn, valueColumn, includeHeaders=True):
    import csv
    # Instatiate an empty dictionary
    d = dict()

    with open(table, 'rb') as csvF:
        csvReader = csv.reader(csvF)
        # Read the first row
        headers = csvReader.next()

        if type(keyColumn) == int:
            keyHeader = headers[keyColumn]
            iKey = keyColumn
        else:
            keyHeader = keyColumn
            # Identify the indices for the key and value fields
            iKey = headers.index(keyColumn)

        if type(valueColumn) == int:
            iValue = valueColumn
        else:
            iValue = headers.index(valueColumn)

        # For each row
        for row in csvReader:
            # Get the values from the proper columns
            k = row[iKey]
            v = row[iValue]
            # If the key value is a float, it can't be used as a dictionary key, so
            # convert it to a string
            if type(k) == float:
                k = str(k)
            # Create the dictionary entry
            d[k] = v
        if not includeHeaders:
            del d[keyHeader]
    return d


# Private function to create a dictionary from the passed table and the passed
# key and value fields for an Excel file
def __DictionaryFromColumns_excel(table, keyColumn, valueColumn):
    try:
        import openpyxl
    except:
        print 'To perform the function DictionaryFromColumns() on an Excel file, you must have the openpyxl library installed.'
        return False

    if not type(keyColumn) == int:
        # Call the function to get the index of the fields
        iKey = GetColumnIndex_excel(table, keyColumn)
    else:
        iKey = keyColumn

    if not type(valueColumn) == int:
        iVal = GetColumnIndex_excel(table, valueColumn)
    else:
        iVal = valueColumn

    # Create the workbook and worksheet objects
    wb = openpyxl.load_workbook(filename = table, use_iterators = True)
    wsName = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(name = wsName)

    # Instantiate an empty dictionary
    d = dict()

    # Instantiate a row count (will be used to skip the first row)
    count = 0
    # For each row in the worksheet
    for row in ws.iter_rows():
        # If the row count is less than one
        if count < 1:
            # Increment the row count and skip to the next row
            count += 1
            continue
        # Get the row's values
        k = row[iKey].internal_value
        v = row[iVal].internal_value
        # If the key value is a float, it can't be used as a dictionary key, so
        # convert it to a string
        if type(k) == float:
            k = str(k)
        # Add the key and value to the dictionary
        d[k] = v

    return d


# Private function to create a dictionary from the passed table and the passed
# key and value fields for a txt file
def __DictionaryFromColumns_txt(document, keyColumn, valueColumn, includeHeaders=True):
    # Create an empty dictionary
    d = {}

    delimiter = DetermineDelimiter(document)

    # If no legitimate delimiter was found, return None
    if delimiter is False:
        print "Could not determine the delimiter used for %s." % document
        return d

    cont = open(document, 'r')
    try:
        # If the user entered text for either of the column identifiers,
        # get the index of the column matching that text.
        rowList = cont.readline().strip().split(delimiter)
        if type(keyColumn) is str:
            keyHeader = keyColumn
            keyColumn = rowList.index(keyColumn)
        else:
            keyHeader = rowList[keyColumn]
        if type(valueColumn) is str:
            valueColumn = rowList.index(valueColumn)

        # Reset the cursor in the file
        cont.seek(0)

        # For each line in the table
        for line in cont.readlines():
            line = line.strip()
            # Create a list of the fields
            cells = line.split(delimiter)
            # Get the key and value content
            key = cells[keyColumn].strip()
            val = cells[valueColumn].strip()
            # Create a new dictionary entry
            d[key] = val

        if not includeHeaders:
            del d[keyHeader]

    except Exception as e:
        print e

    # Close the document on the way out of the function.
    finally:
        cont.close()
        # Return the dictionary
        return d




###################################################
########### DICTIONARY - FULL TABLE ###############
###################################################

##################################
#### Public function to create a dictionary from a table
def DictionaryFromTable(document, keyColumn, keyWildcard="", delimiter="", includeHeaders=True):
    '''
    (string, string/integer, string/integer, [string], [string], [boolean]) -> dictionary

    Returns a dictionary from a table. The key is the content from the column
    submitted as the second argument; the value is a tuple of the content from
    all other columns in the table.

    Note: If you wish for the dictionary's values to contain the information
    from only one column in the table, use the DictionaryFromColumns() function.

    Arguments:
    document -- The full, absolute path to the table file.
    keyColumn -- The header text or 0-based index of the column that will be
        assigned as the dictionary key.
    keyWildcard -- An optional argument to specify a string of text that must
        appear in the key field for that row to be included in the dictionary.
    delimiter -- An optional argument to specify the table's delimiting
        character/string. If none entered, the script will attempt to determine
        the delimiter on its own.
    includeHeaders -- An optional, boolean parameter to indicate whether you
        wish for the table's headers (the first row in the table) to be included.
        By default, it is set to True, meaning that headers will be included
        in the returned dictionary.

    Examples:
    >>> DictionaryFromTable('test.txt', 1)
    {'eng': ('num', 'span'), 'three': ('3', 'tres'), 'one': ('1', 'uno'), 'four': ('4', 'quatro'), 'five': ('5', 'cinco'), 'two': ('2', 'dos')}
    >>> DictionaryFromTable('test.txt', 2, 'o')
    {'dos': ('2', 'two'), 'cinco': ('5', 'five'), 'quatro': ('4', 'four'), 'uno': ('1', 'one')}
    '''
    # Create an empty dictionary
    d = {}

    # If the user did not enter a delimiter,
    if delimiter == '':
        # Get the delimiter
        delimiter = DetermineDelimiter(document)
    # If the user did enter a delimiter, use it
    else:
        delm = delimiter

    # If no legitimate delimiter was found, return None
    if delimiter == False:
        print "Could not determine the delimiter used for %s." % document
        return d

    cont = open(document, 'r')
    try:
        # If the user entered text for either of the column identifiers,
        # get the index of the column matching that text.
        rowList = cont.readline().strip().split(delimiter)
        if type(keyColumn) is str:
            keyHeader = keyColumn
            keyColumn = rowList.index(keyColumn)
        else:
            keyHeader = rowList[keyColumn]

        # Reset the cursor in the file
        cont.seek(0)

        # For each line in the table
        for line in cont.readlines():
            line = line.strip()
            # Create a list of the fields
            cells = line.split(delimiter)
            # Get the key and value content
            keyC = keyColumn
            key = cells[keyC].strip()
            del cells[keyC]
            tup = tuple(cells)
            # If the user set wildcards, only continue if the key/value are valid
            if keyWildcard in key:
                # Create a new dictionary entry
                d[key] = tup

        if not includeHeaders:
            del d[keyHeader]

    except Exception as e:
        print e

    # Close the document on the way out of the function.
    finally:
        cont.close()
        # Return the dictionary
        return d



###################################################
############## LIST FROM TABLE ####################
###################################################

# Public function to create a list of of lists from a table
def ListFromTable(table, includeHeaders=True, delimiter='\t'):
    '''
    (str) -> list

    Returns all the records from the table as a list of lists.

    Argument:
    table -- File from which you wish to return the records
    includeHeaders -- An optional, boolean parameter to indicate whether you
         wish to include the table headers (assumed to be the first row in csv,
         txt, and Excel files) as the first nested list. By default, it is set
         to True, meaning that the headers will be included.

    Example:
    >>> ListFromTable('testTable.csv')
    [['001','bbaeax','Bald Eagle'],['002','mnarox','River Otter']]
    '''
    # Identify the file type and call the appropriate function for the type
    if table.endswith('.csv'):
        l = __ListFromTable_csv(table)
    elif table.endswith('txt'):
        l = __ListFromTable_txt(table, delimiter, includeHeaders)
    elif table.endswith('.xlsx') or table.endswith('.xls'):
        l = __ListFromTable_excel(table)
    elif table.endswith('.shp') or '.gdb' in os.path.abspath(table):
        l = featureclasses.ListFromTable(table)
    else:
        import arcpy
        if '.gdb' in arcpy.env.workspace:
            l = featureclasses.ListFromTable(table)

    if not includeHeaders:
        l = l[1:]

    return l



def __ListFromTable_csv(table):
    l = list()
    with open(table, 'rb') as csvFile:
        reader = csv.reader(csvFile)
        for row in reader:
            l.append(row)
    return l


def __ListFromTable_excel(table):
    # Attempt to import the necesssary library
    try:
        import openpyxl
    except:
        print 'To perform the function ListValues() on an Excel file, you must have the openpyxl library installed.'
        return False

    l = list()
    wb = openpyxl.load_workbook(filename = table, use_iterators = True)
    wsName = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(name = wsName)

    for row in ws.iter_rows():
        rowList = list()
        for cell in row:
            rowList.append(cell.internal_value)
        l.append(rowList)

    return l


def __ListFromTable_txt(document, delimiter, ignoreFirstRow=False):
    '''
    (string, [string]) -> list

    Creates a list of lists from the passed table.

    Arguments:
    document -- The full, absolute path to the table file.
    ignoreFirstRow -- An optional, boolean argument, indicating whether you wish
        include the first row in the returned list. By default, it is set to
        False, meaning that the first row will be included.
    '''
    if not delimiter:
        delimiter = DetermineDelimiter(document)
    # If no legitimate delimiter was found, return None
    if delimiter == False:
        print "Could not determine the delimiter used for %s." % document
        return False

    rows = []

    with open(document, 'r') as cont:
        # For each line in the table
        for line in cont.readlines():
            line = line.strip()
            # Create a list of the fields
            cells = line.split(delimiter)
            rows.append(cells)

    if ignoreFirstRow:
        rows = rows[1:]

    return rows




###################################################
################# GAP TABLES ######################
###################################################

##################################
#### Function to create a table with the full taxonomic, crosswalking, and
#### GAP range/model status information for a list of species.
def WriteSppTable(outTable, spp, delimiter=','):
    '''
    (string, list, [string]) -> string

    Creates a table of the full taxonomic, crosswalking, and range/model status
        info for all the species in the passed list.

    Arguments:
    outTable -- The file to which you wish to write the table.
    spp -- A list of GAP species codes
    delimiter -- An optional parameter to indicate the character(s) with which
        you wish to delimit the table. The default value is ",".

    Example:
    >>> WriteSppTable(r'MyTable.txt', ['mNAROx', 'bbaeax', 'ABOTOX'], '\t')
    '''
    try:

        # List of column headings to be used
        headings = ['GAPCode','Class','Order','Family','Genus','Species','Subspecies','Binomial','CommonName','ELCODE','ITS_TSN','NatureServeID','Published_Range','Published_Model','GAP_National_Download']

        # Create the table and write the headings to it
        docs.Write(outTable, delimiter.join(headings), 'overwrite')

        # For each passed species...
        for sp in spp:
            # Get the species' combined info
            comb = gapdb.SpInfoCombined(sp)
            # If the info is properly returned...
            if type(comb) is tuple or type(comb) is list:
                # Create a list of the values
                combList = [i for i in comb]
                # Add the download name to the list
                combList.append(gapdb.NameDownload(sp))
                # Create a string with the values joined by the delimiter
                sComb = delimiter.join(combList)
                # Write the species' info to the table
                docs.Write(outTable, sComb)
            else:
                pass
        # Return the table
        return outTable

    except Exception as e:
        print e
        return False




###################################################
############### WRITE TABLES ######################
###################################################

def WriteDictToTable(dictionary, outputFile, delimiter=',', headerKey=False, headerValue=False, overwrite=True):
    '''
    (dict, str, [str], [str], [str], [boolean])

    Writes the passed dictionary to a table.

    Arguments:
    dictionary -- The dictionary from which you wish to extract values to write.

    outputFile -- The path to/name of the file to which you wish to write the
        table.

    delimiter -- An optional parameter indicating the delimiter you wish to use
        in the output table. By default, it is set as a comma.

    headerKey -- An optional parameter through which you can pass a header to
        to write to the output table for the dictionary's keys.

    valueKey -- An optional parameter through which you can pass a header to
        to write to the output table for the dictionary's values.

    overwrite -- An optional parameter indicating whether you wish to overwrite
        existing records that may exist in the passed outputFile. By default, it
        is set to true, meaning that the files will be completely overwritten.
    '''
    if overwrite:
        wa = 'wb'
    else:
        wa = 'ab'

    if delimiter == ',' and outputFile.lower().endswith('.csv'):
        try:
            import csv
            with open(outputFile, wa) as csvFile:
                csvWriter = csv.writer(csvFile)
                if headerKey and headerValue and overwrite:
                    csvWriter.writerow([headerKey, headerValue])
                keys = dictionary.keys()
                outList = [[i, dictionary[i]] for i in keys]
                for pair in outList:
                    csvWriter.writerow(pair)
            return outputFile
        except:
            pass

    with open(outputFile, wa) as oF:
        if overwrite and headerKey and headerValue:
            headers = '{0},{1}'.format(headerKey, headerValue)
            oF.write(headers + '\n')
        for i, j in dictionary.iteritems():
            outText = '{0}{1}{2}'.format(i, delimiter, j)
            oF.write(outText + '\n')

    return outputFile



def AttributeTableToCsv(layer, outputFile, keepOID=True, keepShape=False):
    '''
    (str, str, [bool], [bool]) -> str

    Exports the attribute table from the passed shapefile, feature class, raster,
        or geodatabase table to a csv. Returns the output csv path.

    Arguments:
    layer -- The path to the shapefile, feature class, raster, or geodatabase
       table, the attribute table of which will be written to a csv
    outputFile -- The path/name of the csv table you wish to create. Note that
       if a file already exists at that location, the rows (inclusive of the
       field names) will be appended to the existing table.
    keepOID -- An optional, boolean parameter, indicating whether you wish for
       each record's Object ID to be included in the output table. By default,
       it is set to True, meaning that the Object ID will be included.
    keepShape -- An optional, boolean parameter, indicating whether you wish for
       each record's shape object to be included in the output. By default, it
       is set to False, meaning that the shape information will not be included.
       Note that for polygons and/or lines, the shape objects can be quite
       large, as they contain the coordinates of every vertex.

    '''
    rows = ListFromTable(layer, True)

    if not keepOID:
        import arcpy
        d = arcpy.Describe(layer)
        if d.hasOID:
            ofn = d.OIDFieldName
            rows = __RemoveField(rows, ofn)

    if not keepShape:
        rows = __RemoveField(rows, 'Shape')

    outputFile = WriteListToCsv(rows, outputFile)

    return outputFile


# Private function to identify the index of the passed "fieldName" in the
# first item in the passed list and then to remove the item at that index
# from each item in the list. If the 'fieldName' is not in the list's first
# item, then the original list will be returned.
def __RemoveField(rows, fieldName):

    try:
        ind = rows[0].index(fieldName)
    except:
        return rows

    outRows = list()

    for row in rows:
        del row[ind]
        outRows.append(row)

    return outRows



def WriteListToCsv(inputList, outputFile, headers=False):
    '''
    (list/tuple, str, [list]) -> str

    Write a list or a list of lists to a csv.

    Parameters:

    inputList -- A list containing the content to be written to the document.
        If the input list is a list of lists (or a tuple of tuples), each
        sub-list will be written as its own row. If the input list is of strings
        and/or numbers, the entire input list will be written as a single row.
    outputFile -- The path to the csv file to which the data will be written.
    headers -- An optional parameter containing the column headings to be
        written to the table. By default, no headings are required, but the
        first sublist in the input list can be treated as the heading row.
    '''
    # On 9/9/2014, changed mode from 'wb', so it will no longer overwrite files.
    if not outputFile.lower().endswith('.csv'):
        outputFile = '{0}.csv'.format(outputFile)

    with open(outputFile, 'ab') as csvFile:
        csvWriter = csv.writer(csvFile)
        if headers:
            csvWriter.writerow(headers)
        test = inputList[0]
        if type(test) == list or type(test) == tuple:
            for row in inputList:
                csvWriter.writerow(row)
        else:
            csvWriter.writerow(inputList)
    return outputFile


##################################
#### Module's main function
def __main():
    pass

# If the module was run directly, call the main function
if __name__ == '__main__':
    __main()
